# app.py
import io
import re
from datetime import datetime, date
from typing import List, Tuple, Optional, Dict
import pandas as pd
import openpyxl
from openpyxl.styles import numbers
import streamlit as st
import os

st.set_page_config(page_title="sachnv_phieuxuatnhap", layout="wide")
# Xóa cache cũ (nếu cần)
st.cache_data.clear()
st.cache_resource.clear()
DATE_FMT_OUT = "%d-%m-%Y"  # dùng cho hiển thị; khi ghi Excel sẽ set number_format

# ===================== Helpers =====================
# ===== PN helpers & expanders (đÃ chuẩn hóa theo yêu cầu) =====
def safe_restart(reason="unknown"):
    """Restart app nhẹ khi gặp lỗi nặng hoặc MemoryError"""
    st.warning(f"⚠️ Ứng dụng gặp lỗi ({reason}). App sẽ khởi động lại trong giây lát...")
    st.experimental_rerun()

def _replace_tail_full(base_full: int, end_token: str) -> int:
    """Thay *toàn bộ* len(end_token) chữ số cuối của base_full bằng end_token."""
    base_str = str(base_full)
    k = min(len(end_token), len(base_str))
    return int(base_str[:-k] + end_token[-k:].zfill(k))

def pn_expand_range_pattern(base_full: int, end_token: str, count: int) -> List[int]:
    """
    BASE…END(k): sinh 'count' số liên tiếp bắt đầu từ số nhỏ hơn giữa BASE và BASE(thay END).
    """
    candidate = _replace_tail_full(base_full, end_token)
    start = min(base_full, candidate)
    return list(range(start, start + count))

def as_date(d):
    """Chuẩn hoá giá trị ngày từ Excel về datetime.date hoặc None."""
    if isinstance(d, datetime): return d.date()
    if isinstance(d, date):     return d
    # Chuỗi dd/mm/yyyy hoặc dd-mm-yyyy
    if isinstance(d, str):
        s = d.strip().replace("-", "/")
        m = re.fullmatch(r"(\d{2})/(\d{2})/(\d{4})", s)
        if m:
            return datetime.strptime(s, "%d/%m/%Y").date()
    return None

def write_excel_with_formats(df: pd.DataFrame, file_name: str, sheet_name: str,
                             ticket_col: str, date_cols: List[str]):
    """Ghi DataFrame ra Excel, ép định dạng:
       - ticket_col: Text '@'
       - date_cols: Date format dd-mm-yyyy.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ghi tạm dữ liệu
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        # xác định cột
        headers = {cell.value: cell.col_idx for cell in ws[1]}
        # ticket -> text
        if ticket_col in headers:
            col = headers[ticket_col]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = numbers.FORMAT_TEXT  # '@'
        # date -> dd-mm-yyyy
        for dc in date_cols:
            if dc in headers:
                col = headers[dc]
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, col).number_format = "DD-MM-YYYY"
    output.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    st.download_button(
        f"Tải {file_name}",
        data=output.getvalue(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ===================== PX: parser =====================
def parse_header_ngay(text: str) -> Tuple[Optional[date], Optional[date]]:
    """'Ngày dd/mm/yyyy (bàn giao ngày dd/mm[/yyyy])' → (ngày xuất, ngày bàn giao)."""
    if not text: return None, None
    m1 = re.search(r"Ngày\s+(\d{2}/\d{2}/\d{4}).*?bàn giao ngày\s+(\d{2}/\d{2})(?:/(\d{4}))?", text, flags=re.I)
    if not m1:
        m1 = re.search(r"Ngày\s+(\d{2}/\d{2}/\d{4}).*?(\d{2}/\d{2})(?:/(\d{4}))?", text, flags=re.I)
    if m1:
        nx = datetime.strptime(m1.group(1), "%d/%m/%Y").date()
        gy = int(m1.group(3)) if m1.group(3) else nx.year
        bg = datetime.strptime(f"{m1.group(2)}/{gy}", "%d/%m/%Y").date()
        return nx, bg
    m2 = re.search(r"(\d{2}/\d{2}/\d{4})", text)
    if m2:
        nx = datetime.strptime(m2.group(1), "%d/%m/%Y").date()
        return nx, nx
    return None, None

def extract_px_tickets_from_row(values: List[str]) -> List[str]:
    out = []
    for v in values:
        if v is None: continue
        s = str(v).strip()
        if not s: continue
        # nhiều cụm trong 1 ô
        tokens = re.findall(r"\b\d{4,}\s*-\s*[\dA-Za-z]+", s)
        if tokens:
            for token in tokens:
                first = token.split("-", 1)[0].strip()
                if first.isdigit(): out.append(first)
        elif "-" in s:
            first = s.split("-", 1)[0].strip()
            if first.isdigit(): out.append(first)
    return out

def parse_px_sheet(ws) -> pd.DataFrame:
    date_rows = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().startswith("Ngày"):
            date_rows.append(r)
    if not date_rows:
        return pd.DataFrame(columns=["Mã CT","Phiếu xuất","Ngày xuất","Ngày bàn giao"])
    date_rows.append(ws.max_row + 1)

    out_rows = []
    for i in range(len(date_rows) - 1):
        start, end = date_rows[i], date_rows[i + 1] - 1
        header = (ws.cell(start, 1).value or "").strip()
        nx, bg = parse_header_ngay(header)
        if not nx: continue
        for r in range(start + 1, end + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for tk in extract_px_tickets_from_row(row_vals):
                out_rows.append({
                    "Mã CT": "PX",
                    "Phiếu xuất": str(tk),      # sẽ định dạng Text khi ghi file
                    "Ngày xuất": as_date(nx),
                    "Ngày bàn giao": as_date(bg),
                })
    return pd.DataFrame(out_rows)

# ===================== PN: parser (đã mở rộng luật) =====================
# ===== PN: mở rộng luật (đồng bộ tham số left_full) =====
def _replace_tail(prev_num: int, sfx: str) -> int:
    """Thay đúng số chữ số cuối của prev_num bằng sfx (dùng cho danh sách hậu tố)."""
    k = len(sfx)
    mod = 10 ** k
    return prev_num - (prev_num % mod) + int(sfx)

def pn_expand_rhs(rhs: str, left_full: Optional[str]) -> List[int]:
    """
    Phân tích phần bên phải dấu '-':
    - Dạng range: BASE...END(k) hoặc BASE…END(k)
    - Dạng liệt kê: FULL, sfx1, sfx2, ...
    - Dạng rút gọn mạnh: chỉ sfx,sfx,... → dùng left_full làm nền (không thêm vào kết quả)
    """
    rhs = rhs.strip()
    out: List[int] = []

    # ---- Case: range (chấp nhận ... hoặc …)
    m = re.fullmatch(r"(\d+)[.…]{3}(\d+)\((\d+)\)", rhs)
    if m:
        base_full = int(m.group(1))
        end_token = m.group(2)
        count = int(m.group(3))
        return pn_expand_range_pattern(base_full, end_token, count)

    # ---- Case: danh sách hậu tố
    parts = [p.strip() for p in rhs.split(",") if p.strip()]
    if not parts:
        return out

    left_num = int(left_full) if (left_full and str(left_full).isdigit()) else None

    first = parts[0]
    if first.isdigit() and len(first) >= 5:
        prev = int(first)
        out.append(prev)
        iterable = parts[1:]
    else:
        if left_num is None:
            return out
        prev = left_num
        iterable = parts

    for token in iterable:
        if not token.isdigit():
            continue
        k = len(token)
        mod = 10 ** k
        new_num = prev - (prev % mod) + int(token)
        # đảm bảo ≥ 5 chữ số
        if len(str(new_num)) < 5 and left_num is not None:
            mod = 10 ** len(token)
            new_num = left_num - (left_num % mod) + int(token)
        out.append(new_num)
        prev = new_num

    return out

def parse_pn_cell(cell_value: str, want_return_suffix: bool) -> List[str]:
    """
    Tách mọi cụm '<left>-<right>' trong 1 ô, bao gồm cả dạng rút gọn:
    - 138584-56…91(9) → 138556..138564
    - 153502-03…37(10) → 153503..153512

    Bổ sung:
    ✅ Nếu phần sau dấu '-' có >=4 chữ số thì không ghép left (dùng nguyên)
    ✅ Nếu các phần sau dấu ',' có ký tự lạ hoặc khoảng trắng -> loại bỏ, chỉ giữ số
    """
    s = str(cell_value).strip()
    if not s:
        return []
    results: List[str] = []

    # --- Dạng rút gọn: <left>-<short>…<end>(<count>)
    m = re.fullmatch(r"(\d+)-(\d+)[.…]{3}(\d+)\((\d+)\)", s)
    if m:
        left = m.group(1)
        short_part = m.group(2)
        end_token = m.group(3)
        count = int(m.group(4))

        base_full = int(left)
        k_short = len(short_part)
        if k_short <= len(left):
            base_full = int(left[:-k_short] + short_part)
        candidate = _replace_tail_full(base_full, end_token)
        start = min(base_full, candidate)
        nums = list(range(start, start + count))
        results.extend([str(n) for n in nums])

    # --- Dạng bình thường hoặc liệt kê
    for left, right in re.findall(r"(\d+)\s*-\s*([0-9,().'’\.\s]+)", s):
        # bỏ qua dạng rút gọn vừa xử lý
        if re.fullmatch(r"\d+[.…]{3}\d+\(\d+\)", right):
            continue

        # ✅ Làm sạch ký tự lạ
        right = re.sub(r"[^0-9,.\s]", "", right).replace(" ", "")
        parts = [p for p in right.split(",") if p]

        if not parts:
            continue

        first = parts[0]
        use_right_as_base = len(first) >= 4  # >=4 chữ số thì không ghép left

        nums = []
        # ✅ Nếu phần đầu đủ dài (>=4 chữ số), dùng nguyên và sinh tiếp theo
        if use_right_as_base:
            base_num = int(first)
            nums.append(base_num)
            for p in parts[1:]:
                p_digits = re.sub(r"\D", "", p)
                if not p_digits:
                    continue
                # thay đuôi của base bằng phần mới
                next_num = _replace_tail_full(base_num, p_digits)
                nums.append(next_num)
        else:
            # logic cũ (ghép với left)
            nums = pn_expand_rhs(right, left_full=left)

        for n in nums:
            results.append(str(n))

    # --- Loại trùng và thêm hậu tố -R nếu cần
    seen = set()
    uniq = []
    for x in results:
        tag = f"{x}-R" if want_return_suffix else x
        if tag not in seen:
            seen.add(tag)
            uniq.append(tag)
    return uniq


def parse_pn_simple_table(ws) -> pd.DataFrame:
    """
    Trường hợp sheet PN dạng đơn giản (không RETURN, không ghép, không ký tự đặc biệt).
    Hỗ trợ cả tiêu đề HAW (thay cho NGUỒN).
    Đọc cột: SỐ | (NGUỒN hoặc HAW) | NGÀY | NGÀY GIAO
    """
    # Tìm dòng tiêu đề có chữ 'SỐ' hoặc 'SO'
    header_row = None
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 10) + 1)]
        texts = [str(v).strip().upper() if v else "" for v in row_vals]
        if any(t.startswith("SỐ") or t == "SO" for t in texts):
            header_row = r
            break
    if not header_row:
        return pd.DataFrame()

    # Xác định vị trí cột dựa theo tiêu đề
    headers = [str(ws.cell(header_row, c).value).strip().upper() if ws.cell(header_row, c).value else "" 
               for c in range(1, ws.max_column + 1)]

    def find_col(keyword_list):
        for kw in keyword_list:
            for i, t in enumerate(headers):
                if kw in t:
                    return i + 1
        return None

    c_so = find_col(["SỐ", "SO"])
    # ✅ “Nguồn” có thể là “NGUỒN”, “NGUON”, hoặc “HAW”
    c_nguon = find_col(["NGUỒN", "NGUON", "HAW"])
    c_ngay = find_col(["NGÀY", "NGAY"])
    c_ngay_giao = find_col(["NGÀY GIAO", "NGAY GIAO"])

    if not c_so:
        return pd.DataFrame()

    rows = []
    r = header_row + 1
    while r <= ws.max_row:
        so = ws.cell(r, c_so).value
        if so is None or str(so).strip() == "":
            # dừng khi gặp dòng trống
            if all(ws.cell(r, c).value in (None, "") for c in range(1, min(ws.max_column, 6)+1)):
                break
            r += 1
            continue

        so_text = str(int(so)) if isinstance(so, (int, float)) else str(so).strip()
        nguon = ws.cell(r, c_nguon).value if c_nguon else ""
        ngay = as_date(ws.cell(r, c_ngay).value) if c_ngay else None
        ngay_giao = as_date(ws.cell(r, c_ngay_giao).value) if c_ngay_giao else None

        rows.append({
            "Số phiếu gốc": so_text,
            "Mã CT": "PN",
            "Nguồn": str(nguon).strip() if nguon else "",
            "Phiếu nhập": so_text,
            "Ngày nhập": ngay,
            "Ngày bàn giao": ngay_giao
        })
        r += 1

    return pd.DataFrame(rows)


def guess_pn_header(ws) -> Dict[str, int]:
    """
    Xác định dòng tiêu đề và vị trí các cột chính (SỐ PHIẾU, NGUỒN, NGÀY, NGÀY GIAO)
    - Dò chữ linh hoạt (có dấu, không dấu, viết hoa/thường, có khoảng trắng)
    - Áp dụng cho bảng RETURN phức tạp
    """
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 15) + 1)]
        texts = [str(v).strip().upper().replace(" ", "") if v is not None else "" for v in vals]

        # tìm vị trí cột SỐ PHIẾU
        c_phieu = next((i+1 for i,t in enumerate(texts) if "SỐPHIẾU" in t or "SOPHIEU" in t), None)
        if not c_phieu:
            continue

        # cột NGUỒN (dò rộng: NGUON, NGUỒN, NG)
        c_nguon = next(
            (i+1 for i,t in enumerate(texts) 
             if "NGUỒN" in t or "NGUON" in t or re.fullmatch(r"NG", t)), 
            None
        )

        # cột NGÀY (phải loại trừ “NGÀY GIAO”)
        c_ngay = next(
            (i+1 for i,t in enumerate(texts)
             if ("NGÀY" in t or "NGAY" in t) and "GIAO" not in t),
            None
        )

        # cột NGÀY GIAO
        c_giao = next(
            (i+1 for i,t in enumerate(texts)
             if "NGÀYGIAO" in t or "NGAYGIAO" in t),
            None
        )

        return {"row": r, "so_phieu": c_phieu, "nguon": c_nguon, "ngay": c_ngay, "ngay_giao": c_giao}
    return {}


def sheet_has_return_flag(ws) -> bool:
    """Nếu vùng tiêu đề (vài dòng đầu) có chữ RETURN → True."""
    for r in range(1, min(10, ws.max_row) + 1):
        for c in range(1, min(6, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "RETURN" in v.upper():
                return True
    return False

def parse_pn_sheet(ws) -> pd.DataFrame:
    """
    Xử lý sheet PN:
    - Nếu có RETURN hoặc có ký tự đặc biệt ('...', ',', '(') -> dạng RETURN/phức tạp
    - Ngược lại -> dạng SIMPLE (bảng tháng)
    Tự động đọc đúng cột NGUỒN kể cả khi merge hoặc công thức.
    """
    # 1️⃣ Kiểm tra xem sheet có RETURN trong tiêu đề hay không
    want_R = sheet_has_return_flag(ws)

    # 2️⃣ Kiểm tra xem cột đầu tiên có chứa ký tự đặc biệt ('-', ',', '.', '(')
    #    để nhận diện dạng phức tạp
    complex_found = False
    for r in range(1, min(ws.max_row, 10)):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and any(x in v for x in ["-", ",", ".", "(", ")"]):
            complex_found = True
            break

    # 3️⃣ Nếu không có RETURN và không có ký tự đặc biệt -> dạng SIMPLE
    if not want_R and not complex_found:
        df_simple = parse_pn_simple_table(ws)
        if not df_simple.empty:
            return df_simple

    # 4️⃣ Còn lại là dạng RETURN / phức tạp
    pos = guess_pn_header(ws)
    if not pos:
        return pd.DataFrame(columns=["File","Sheet","Số phiếu gốc","Mã CT","Nguồn","Phiếu nhập","Ngày nhập","Ngày bàn giao"])

    r0 = pos["row"] + 1
    c_phieu, c_nguon, c_ngay, c_giao = pos["so_phieu"], pos["nguon"], pos["ngay"], pos["ngay_giao"]

    # nếu không dò được cột NGUỒN, mặc định ngay sau SỐ PHIẾU
    if not c_nguon:
        c_nguon = c_phieu + 1

    out_rows = []
    r = r0
    while r <= ws.max_row:
        raw = ws.cell(r, c_phieu).value
        if raw is None or str(raw).strip() == "":
            if all(ws.cell(r, c).value in (None, "") for c in range(1, min(ws.max_column, 6)+1)):
                break
            r += 1
            continue

        raw_str = str(raw).strip()          # SỐ PHIẾU GỐC
        # đọc cột NGUỒN: nếu bị merge/công thức, dùng internal_value fallback
        cell_nguon = ws.cell(r, c_nguon)
        nguon_val = cell_nguon.value or getattr(cell_nguon, "internal_value", None)
        nguon = str(nguon_val).strip() if nguon_val else ""

        ngay  = as_date(ws.cell(r, c_ngay).value)  if c_ngay  else None
        giao  = as_date(ws.cell(r, c_giao).value)  if c_giao  else None

        nums = parse_pn_cell(raw_str, want_return_suffix=want_R)
        for phieu in nums:
            out_rows.append({
                "Số phiếu gốc": raw_str,
                "Mã CT": "PN",
                "Nguồn": nguon,
                "Phiếu nhập": phieu,
                "Ngày nhập": ngay,
                "Ngày bàn giao": giao,
            })
        r += 1

    return pd.DataFrame(out_rows)

try:
    # ===================== UI ================================
    st.title("KIỂM TRA PHIẾU XUẤT - NHẬP - TỒN")

    st.markdown("""
    **Bước 1.** Upload **một hoặc nhiều** file Excel.  
    **Bước 2.** Chọn đúng **sheet** cần xử lý trong mỗi file.  
    **Bước 3.** Chọn chế độ **Xử lý phiếu xuất (PX)** hoặc **Xử lý phiếu nhập (PN)** → bấm **Xử lý** để xử lý file.
    **Bước 4.** liên hệ nguyenvansach báo lỗi (nếu cần).
    """)

    uploaded_files = st.file_uploader("Chọn file Excel", type=["xlsx", "xlsm"], accept_multiple_files=True)

    if not uploaded_files:
        st.info("Hãy tải lên ít nhất một file Excel.")
        st.stop()

    # load workbooks
    workbooks = {}
    file_sheets = {}
    for f in uploaded_files:
        bio = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(bio, data_only=True)
        workbooks[f.name] = wb
        file_sheets[f.name] = wb.sheetnames
        wb.close()
    st.write("### Chọn sheet để xử lý")
    selected_sheets = {}
    cols = st.columns(min(3, len(file_sheets)))
    for i, (fname, sheets) in enumerate(file_sheets.items()):
        with cols[i % len(cols)]:
            st.caption(f"**{fname}**")
            selected = st.multiselect(f"Sheet trong {fname}", sheets, default=sheets, key=f"ms_{fname}")
            selected_sheets[fname] = selected

    mode = st.radio("Chọn loại phiếu cần xử lý", options=["PX","PN"], horizontal=True)

    if mode == "PX":
        if st.button("Xử lý dữ liệu phiếu xuất", type="primary"):
            all_rows = []
            for fname, sheets in selected_sheets.items():
                wb = workbooks[fname]
                for sheet in sheets:
                    ws = wb[sheet]
                    df = parse_px_sheet(ws)
                    if not df.empty:
                        df.insert(0, "File", fname)
                        df.insert(1, "Sheet", sheet)
                        all_rows.append(df)
            if not all_rows:
                st.warning("Không trích xuất được dữ liệu phiếu xuất từ các sheet đã chọn.")
            else:
                df_all = pd.concat(all_rows, ignore_index=True)
                st.success(f"Đã trích xuất {len(df_all)} dòng PX.")
                st.dataframe(df_all.head(200).assign(
                    **{"Ngày xuất": df_all["Ngày xuất"].map(lambda d: d.strftime(DATE_FMT_OUT) if d else ""),
                    "Ngày bàn giao": df_all["Ngày bàn giao"].map(lambda d: d.strftime(DATE_FMT_OUT) if d else "")}
                ))
                # ✅ Lưu dữ liệu PX để dùng cho phần so sánh BRAVO
                # Lưu phiên bản rút gọn (chỉ cột cần so sánh)
                cols_need = ["Mã CT", "Phiếu nhập", "Phiếu xuất"]
                cols_exist = [c for c in cols_need if c in df_all.columns]
                st.session_state.last_merged = df_all[cols_exist].copy()
                timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%M")
                write_excel_with_formats(
                    df_all, file_name=f"PX_raw_output_{timestamp}.xlsx", sheet_name="PX_raw",
                    ticket_col="Phiếu xuất", date_cols=["Ngày xuất", "Ngày bàn giao"]
                )

    else:  # PN
        if st.button("Xử lý dữ liệu phiếu nhập", type="primary"):
            all_rows = []
            for fname, sheets in selected_sheets.items():
                wb = workbooks[fname]
                for sheet in sheets:
                    ws = wb[sheet]
                    df = parse_pn_sheet(ws)
                    if not df.empty:
                        df.insert(0, "File", fname)
                        df.insert(1, "Sheet", sheet)
                        all_rows.append(df)
            if not all_rows:
                st.warning("Không trích xuất được dữ liệu phiếu nhập từ các sheet đã chọn.")
            else:
                df_all = pd.concat(all_rows, ignore_index=True)
                # ... trong nhánh if mode == "PN": sau khi df_all = pd.concat(...)
                order = ["File","Sheet","Số phiếu gốc","Mã CT","Nguồn","Phiếu nhập","Ngày nhập","Ngày bàn giao"]
                for col in order:
                    if col not in df_all.columns:
                        df_all[col] = ""    # phòng khi sheet nào đó thiếu
                df_all = df_all.reindex(columns=order)

                st.success(f"Đã trích xuất {len(df_all)} dòng PN.")
                st.dataframe(df_all.head(200).assign(
                    **{"Ngày nhập": df_all["Ngày nhập"].map(lambda d: d.strftime(DATE_FMT_OUT) if d else ""),
                    "Ngày bàn giao": df_all["Ngày bàn giao"].map(lambda d: d.strftime(DATE_FMT_OUT) if d else "")}
                ))
                # ✅ Lưu dữ liệu GHEP (PN) vào session để dùng cho bước so sánh BRAVO
                # Lưu phiên bản rút gọn (chỉ cột cần so sánh)
                cols_need = ["Mã CT", "Phiếu nhập", "Phiếu xuất"]
                cols_exist = [c for c in cols_need if c in df_all.columns]
                st.session_state.last_merged = df_all[cols_exist].copy()
                # Ghi file: cột phiếu dạng Text, ngày dạng Date dd-mm-yyyy
                timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%M")
                write_excel_with_formats(
                    df_all,
                    file_name=f"PN_raw_output_{timestamp}.xlsx",
                    sheet_name="PN_raw",
                    ticket_col="Phiếu nhập",
                    date_cols=["Ngày nhập", "Ngày bàn giao"]
                )
    import pandas as pd
    from openpyxl.styles import PatternFill

    st.markdown("---")
    st.header("🔍 So sánh số phiếu với số phiếu trên BRAVO")

    uploaded_bravo = st.file_uploader(
        "Tải các file xuất ra từ BRAVO để so sánh (Đảm bảo cột A=Mã, cột C=Số phiếu)",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True,
        key="bravo"
    )

    if "last_merged" not in st.session_state:
        st.session_state.last_merged = None

    if uploaded_bravo:
        st.write("### Chọn sheet cần so sánh trong từng file:")
        selected_sheets_bravo = {}
        cols = st.columns(min(3, len(uploaded_bravo)))
        for i, f in enumerate(uploaded_bravo):
            wb = openpyxl.load_workbook(f, read_only=True)
            with cols[i % len(cols)]:
                st.caption(f"**{f.name}**")
                selected = st.multiselect(
                    f"Sheet trong {f.name}",
                    wb.sheetnames,
                    default=wb.sheetnames,
                    key=f"bravo_{f.name}"
                )
                selected_sheets_bravo[f.name] = selected
            wb.close()
    if uploaded_bravo and st.session_state.last_merged is not None:
        btn_compare = st.button("⚖️ So sánh với file BRAVO")
        if btn_compare:
            # --- Đọc dữ liệu BRAVO ---
            all_bravo = []
            for f in uploaded_bravo:
                wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
                for sheet in selected_sheets_bravo.get(f.name, wb.sheetnames):
                    ws = wb[sheet]
                    # Tìm tiêu đề có "Mã" và "Số"
                    header_row = None
                    for r in range(1, 10):
                        row_vals = [ws.cell(r, c).value for c in range(1, 6)]
                        if any(str(v).strip().upper() == "MÃ" for v in row_vals if v):
                            header_row = r
                            break
                    if not header_row:
                        continue
                    col_ma = 1  # A
                    col_so = 3  # C
                    rows = []
                    for r in range(header_row + 1, ws.max_row + 1):
                        ma = ws.cell(r, col_ma).value
                        so = ws.cell(r, col_so).value
                        if ma and so:
                            ma = str(ma).strip().upper()
                            # bỏ khoảng trắng trong số
                            so_text = str(so).strip().replace(" ", "")
                            rows.append({"Mã": ma, "Số": so_text})
                    if rows:
                        df_b = pd.DataFrame(rows)
                        df_b["File"] = f.name
                        df_b["Sheet"] = sheet
                        all_bravo.append(df_b)
            if not all_bravo:
                st.warning("❌ Không đọc được dữ liệu hợp lệ từ file BRAVO.")
            else:
                df_bravo = pd.concat(all_bravo, ignore_index=True)
                st.success(f"✅ Đã đọc {len(df_bravo)} dòng từ {len(uploaded_bravo)} file BRAVO.")

                # --- So sánh ---
                df_ghep = st.session_state.last_merged.copy()
                df_ghep["Loại"] = df_ghep["Mã CT"].str.upper().str.strip()
                df_bravo["Loại"] = df_bravo["Mã"].str.upper().str.strip()

                # Chuẩn hóa tên cột so sánh
                col_map = {"PX": "Phiếu xuất", "PN": "Phiếu nhập"}
                result_rows = []
                for loai in ["PX", "PN"]:
                    col_phieu = col_map[loai]
                    df_gh = df_ghep[df_ghep["Loại"] == loai].copy()
                    df_br = df_bravo[df_bravo["Loại"] == loai].copy()
                    bravo_set = set(df_br["Số"].astype(str).str.replace(" ", ""))

                    for _, row in df_gh.iterrows():
                        so_phieu = str(row[col_phieu]).strip().replace(" ", "")
                        co_trong_bravo = so_phieu in bravo_set
                        row_out = row.to_dict()
                        row_out["Trạng thái"] = "" if co_trong_bravo else "⚠️ Bravo không có"
                        result_rows.append(row_out)
                wb.close()
                df_result = pd.DataFrame(result_rows)
                st.dataframe(df_result.head(200))
                st.info(f"Tổng số dòng: {len(df_result)}")
                
                # --- Ghi file Excel với tô màu ---
                out_path = io.BytesIO()
                with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                    df_result.to_excel(writer, index=False, sheet_name="So_sanh")
                    ws = writer.book["So_sanh"]
                    headers = {cell.value: cell.col_idx for cell in ws[1]}
                    col_phieu = headers.get("Phiếu nhập") or headers.get("Phiếu xuất")
                    col_trangthai = headers.get("Trạng thái")

                    yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
                    for r in range(2, ws.max_row + 1):
                        if ws.cell(r, col_trangthai).value == "⚠️ Bravo không có":
                            if col_phieu:
                                ws.cell(r, col_phieu).fill = yellow
                            ws.cell(r, col_trangthai).fill = yellow
                timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%M")
                st.download_button(
                    "⬇️ Tải file kết quả so sánh",
                    data=out_path.getvalue(),
                    file_name=f"So_sanh_PX_PN_vs_BRAVO_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    else:
        st.info("⚙️ Hãy xử lý file nhập/xuât trước, sau đó tải file BRAVO để so sánh.")

except MemoryError:
    safe_restart("thiếu bộ nhớ")

except Exception as e:
    st.error(f"❌ Lỗi không mong muốn: {e}")
    safe_restart("lỗi không xác định")